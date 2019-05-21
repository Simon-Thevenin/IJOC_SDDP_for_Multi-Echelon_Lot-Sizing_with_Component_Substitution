from __future__ import absolute_import, division, print_function
import pandas as pd
from Tool import Tool
import openpyxl as opxl
import cPickle as pickle
from InstanceReaderGrave import InstanceReaderGrave
from InstanceReaderTemplemeier import InstanceReaderTemplemeier
from InstanceReaderJDA import InstanceReaderJDA
from Constants import Constants
#import networkx as nx


#import matplotlib.pyplot as plt

class Instance(object):

    # Constructor
    def __init__(self):
        self.InstanceName = ""
        self.NrProduct = -1
        self.NrTimeBucket = -1
        self.NrResource = -1
        self.Distribution = ""
        self.NrAlternates = -1
        self.Alternates = []

        # The two attributes below define the number of periods where the demand is known with certainty at the
        # beginning and at the end of the horizon
        self.NrTimeBucketWithoutUncertaintyAfter = -1
        self.NrTimeBucketWithoutUncertaintyBefore = -1
        self.NrTimeBucketWithoutUncertainty = -1
        self.LostSaleCost = []
        self.Gamma = 0  # Discount Factor
        self.ProductSet = []
        self.ProductWithExternalDemand = []

        # The table below give an index to each product with an external demand.
        self.ProductWithExternalDemandIndex = []
        self.ProductWithoutExternalDemand = []
        self.ProductWithoutExternalDemandIndex = []
        self.TimeBucketSet = []
        self.ResourceSet = []
        self.ProcessingTime = []
        self.Capacity = []
        self.LeadTimes = []

        # StartingInventories gives the stock level at the start of horizon, and Delivery indicates the planned
        # deliveries which will arrive during the horizon
        self.StartingInventories = []
        self.Delivery = []

        # self.Requirements[q][p] indicates the quantity of p required for production of 1 unit of q
        self.Requirements = []
        self.InventoryCosts = []
        self.SetupCosts = []
        self.BackorderCosts = []
        self.VariableCost = []
        # AternateCosts[p][c] the cost (per unit) of using component c for product p
        self.AternateCosts = []

        # The data below are used to generate the demand distribution
        self.YearlyAverageDemand = []
        self.ForecastedAverageDemand = []
        self.YearlyStandardDevDemands = []
        self.ForecastError = -1
        self.RateOfKnownDemand = -1
        self.ForcastedStandardDeviation = []

        # This variable is true if the end of horizon in the instance is the actual end of horizon
        # (False in a rolling horizon framework)
        self.ActualEndOfHorizon = True

        # The names of the products are used for printing and display purpose only
        self.ProductName = []
        self.ResourceName = []
        ###################################################################################################
        # The data below are not given with the instance, but computed as they help to build the model
        ###################################################################################################
        # True if product p has external demand, and false otherwise
        self.HasExternalDemand = []
        # The set of product which are required for production of each product.
        self.RequieredProduct = []
        # The total requirement include all items (direct components and sub-components) which compose an item
        self.TotalRequirement = []
        # The number of level in the bom
        self.NrLevel = -1
        # The level of each product in the bom
        self.Level = []
        # The number of possible components for each product
        self.NrComponent = []
        # The number of period to produce an item from row material when produced as fast as possible
        self.MaxLeadTimeProduct = []
        # The maximum value in the table MaxLeadTimeProduct (use to be: MaxLeadTime)
        self.MaxMaxLeadTimeProduct = -1
        # the maximum lead time over all items
        self.MaimumLeadTime = -1
        # If this is true, a single scenario with average demand is generated
        # self.BranchingStrategy = 3
        # The maximum quantity of item i which can be produce at each period (used to wave the safety-stock constraints)
        self.MaximumQuanityatT = [[]]

    # This function print the instance on the screen
    def PrintInstance(self):
        print("instance: %s" % self.InstanceName)
        print("instance with %d products and %d time buckets" % (self.NrProduct, self.NrTimeBucket))
        print(
            "requirements: \n %r" % (pd.DataFrame(self.Requirements, index=self.ProductName, columns=self.ProductName)))
        aggregated = [self.LeadTimes, self.StartingInventories, self.InventoryCosts,
                      self.SetupCosts, self.BackorderCosts]
        col = ["Leadtimes", "StartingInventories", "InventoryCosts", "SetupCosts", "BackorderCosts"]
        print("Per product data: \n %r" % (pd.DataFrame(aggregated, columns=self.ProductName, index=col).transpose()))
        print("capacity: \n %r" % (pd.DataFrame(self.Capacity)))
        print("processing time: \n %r" % (pd.DataFrame(self.ProcessingTime, index=self.ProductName)))

    # This function define the current instance as a  small one, used to test the model
    def DefineAsSmallIntance(self):
        self.InstanceName = "SmallIntance"
        self.Distribution = "Normal"
        self.ProductName = ["P1", "P2", "P3", "P4", "P5"]
        self.NrProduct = 5
        self.NrTimeBucket = 6
        self.NrTimeBucketWithoutUncertaintyAfter = 3
        self.NrTimeBucketWithoutUncertaintyBefore = 3
        self.NrResource = 5
        self.Gamma = 0.9
        self.NrAlternates = 6
        self.Alternates = [[0, 1],
                           [2],
                           [3],
                           [4],
                           [5]]
        self.Requirements = [[0, 1, 1, 0, 0],
                             [0, 0, 0, 1, 0],
                             [0, 0, 0, 0, 0],
                             [0, 0, 0, 0, 1],
                             [0, 0, 0, 0, 0]]
        self.LeadTimes = [0, 1, 1, 1, 1]
        self.ProcessingTime = [[1, 0, 0, 0, 0],
                               [0, 2, 0, 0, 0],
                               [0, 0, 5, 0, 0],
                               [0, 0, 0, 1, 2],
                               [0, 0, 0, 1, 5]]
        self.YearlyAverageDemand = [10, 0, 0, 0, 0]
        self.ForecastedAverageDemand = [[10, 0, 0, 0, 0],
                                        [10, 0, 0, 0, 0],
                                        [10, 0, 0, 0, 0],
                                        [10, 0, 0, 0, 0],
                                        [10, 0, 0, 0, 0]]
        self.ForecastError = [0.5, 0, 0, 0, 0]
        self.RateOfKnownDemand = 0.0
        self.YearlyStandardDevDemands = [5, 0, 0, 0, 0]
        self.ForcastedStandardDeviation = [[5, 0, 0, 0, 0],
                                           [5, 0, 0, 0, 0],
                                           [5, 0, 0, 0, 0],
                                           [5, 0, 0, 0, 0],
                                           [5, 0, 0, 0, 0]]
        self.StartingInventories = [10.0, 100.0, 100.0, 100.0, 100.0]
        self.InventoryCosts = [15.0, 4.0, 3.0, 2.0, 1.0]
        self.VariableCost = [4.0, 3.0, 2.0, 1.0, 0.0]
        self.SetupCosts = [10000.0, 1.0, 1.0, 1.0, 1.0]
        self.BackorderCosts = [100000.0, 0.0, 0.0, 0.0, 0.0]  # for now assume no external demand for components
        self.Capacity = [15, 15, 15, 15, 15]
        self.LostSaleCost = [1000.0, 0.0, 0.0, 0.0, 0.0]
        self.ComputeInstanceData()

    # This function defines a very small instance, this is useful for debugging.
    def DefineAsSuperSmallIntance(self):
        self.InstanceName = "SuperSmallIntance"
        self.Distribution = Constants.NonStationary
        self.ProductName = ["P1", "P2a", "P2b"]
        self.NrProduct = 3
        self.NrTimeBucket = 4
        self.NrTimeBucketWithoutUncertaintyAfter = 0
        self.NrTimeBucketWithoutUncertaintyBefore = 2
        self.NrResource = 3
        self.Gamma = 1
        self.Requirements = [[0, 1, 0],
                             [0, 0, 0],
                             [0, 0, 0]]
        self.MasterProduct = [0, 1]
        #Alternates[p][p2] 1 if p2 can be used as alternate for p
        self.Alternates = [[0, 0, 0],
                           [0, 0, 1],
                           [0, 0, 0]]
        #AternateCosts[p][c] the cost (per unit) of using component c for product p
        self.AternateCosts = [[0, 0, 3],
                              [0, 0, 0],
                              [0, 0, 0]]
        self.Capacity = [50, 5, 5]
        self.LeadTimes = [1, 1, 1]
        self.ProcessingTime = [[1, 0, 0],
                               [0, 1, 0],
                               [0, 0, 1]]
        self.YearlyAverageDemand = [10, 0, 0]
        self.ForecastedAverageDemand = [[10, 0, 0],
                                        [10, 0, 0],
                                        [10, 0, 0],
                                        [10, 0, 0],
                                        [10, 0, 0]
                                        ]
        self.ForecastError = [0.25, 0, 0]
        self.RateOfKnownDemand = 0.0
        self.YearlyStandardDevDemands = [2, 0, 0]
        self.ForcastedStandardDeviation = [[2, 0, 0],
                                           [2, 0, 0],
                                           [2, 0, 0],
                                           [2, 0, 0],
                                           [2, 0, 0]]

        self.StartingInventories = [10.0, 5.0, 5.0]
        self.InventoryCosts = [5.0, 2.0, 2.0]
        self.VariableCost = [5.0, 0.0, 0.0]
        self.SetupCosts = [5.0, 5.0, 5.0]
        self.BackorderCosts = [10.0, 0.0, 0.0]  # for now assume no external demand for components
        self.LostSaleCost = [100.0, 0.0, 0.0]
        self.ComputeInstanceData()

        # This function defines a very small instance, this is useful for debugging.
    def DefineAsTwoItemIntance(self):
            self.InstanceName = "TwoItemIntance"
            self.Distribution = Constants.NonStationary
            self.ProductName = ["P1", "P2"]
            self.NrProduct = 2
            self.NrTimeBucket = 2
            self.NrTimeBucketWithoutUncertaintyAfter = 0
            self.NrTimeBucketWithoutUncertaintyBefore = 0
            self.NrResource = 2
            self.Gamma = 1
            self.Requirements = [[0, 1],
                                 [0, 0]]
            self.MasterProduct = [0, 1]
            # Alternates[p][p2] 1 if p2 can be used as alternate for p
            self.Alternates = [[0, 0],
                               [0, 0]]
            # AternateCosts[p][c] the cost (per unit) of using component c for product p
            self.AternateCosts = [[0, 0],
                                  [0, 0]]
            self.Capacity = [50, 50]
            self.LeadTimes = [0, 0]
            self.ProcessingTime = [[1, 0],
                                   [0, 1]]
            self.YearlyAverageDemand = [10, 0]
            self.ForecastedAverageDemand = [[10, 0],
                                            [10, 0],
                                            [10, 0],
                                            [10, 0],
                                            [10, 0]
                                            ]
            self.ForecastError = [0.25, 0]
            self.RateOfKnownDemand = 0.0
            self.YearlyStandardDevDemands = [2, 0]
            self.ForcastedStandardDeviation = [[2, 0],
                                               [2, 0],
                                               [2, 0],
                                               [2, 0],
                                               [2, 0]]

            self.StartingInventories = [10.0, 5.0]
            self.InventoryCosts = [5.0, 2.0]
            self.VariableCost = [5.0, 0.0]
            self.SetupCosts = [5.0, 5.0]
            self.BackorderCosts = [10.0, 0.0]  # for now assume no external demand for components
            self.LostSaleCost = [100.0, 0.0]
            self.ComputeInstanceData()

        # This function defines a very small instance, this is useful for debugging.
    def DefineAsOneItemIntance(self):
            self.InstanceName = "OneItemIntance"
            self.Distribution = Constants.NonStationary
            self.ProductName = ["P1"]
            self.NrProduct = 1
            self.NrTimeBucket = 5
            self.NrTimeBucketWithoutUncertaintyAfter = 0
            self.NrTimeBucketWithoutUncertaintyBefore = 0
            self.NrResource = 1
            self.Gamma = 1
            self.Requirements = [[0]]
            self.MasterProduct = [0]
            # Alternates[p][p2] 1 if p2 can be used as alternate for p
            self.Alternates = [[0]]
            # AternateCosts[p][c] the cost (per unit) of using component c for product p
            self.AternateCosts = [[0]]
            self.Capacity = [50]
            self.LeadTimes = [1]
            self.ProcessingTime = [[1]]
            self.YearlyAverageDemand = [10]
            self.ForecastedAverageDemand = [[10],
                                            [10],
                                            [10],
                                            [10],
                                            [10]
                                            ]
            self.ForecastError = [0.25]
            self.RateOfKnownDemand = 0.0
            self.YearlyStandardDevDemands = [2]
            self.ForcastedStandardDeviation = [[2],
                                               [2],
                                               [2],
                                               [2],
                                               [2]]

            self.StartingInventories = [10.0]
            self.InventoryCosts = [5.0]
            self.VariableCost = [5.0]
            self.SetupCosts = [5.0]
            self.BackorderCosts = [10.0]  # for now assume no external demand for components
            self.LostSaleCost = [100.0]
            self.ComputeInstanceData()

    def IsMaterProduct(self, prod):
        return sum(self.Alternates[p][prod] for p in self.ProductSet) == 0

    def GetMasterProduct(self, prod):
            if sum(self.Requirements[p][prod] for p in self.ProductSet) > 0.0:
                return prod
            else:
                for p in self.ProductSet:
                    if self.Alternates[p][prod]:
                        return p

    # def DrawSupplyChain(self):
    #      G = nx.DiGraph()
    #      labels = {}
    #      for p in self.ProductSet:
    #          G.add_node(p)
    #          labels[p] = self.ProductName[p]
    #          # node.append( G.add_nodes_from(p) )
    #
    #      for p in self.ProductSet:
    #          for q in self.ProductSet:
    #              if self.Requirements[q][p]:
    #                  G.add_edge(p, q, color='b')
    #
    #                  for i in self.ProductSet:
    #                      if self.Alternates[p][i] and p != i:
    #                          G.add_edge(i, q, color='g')
    #
    #      pos = nx.spring_layout(G)
    #
    #      print(pos)
    #      maxl = max(self.Level)
    #      for p in self.ProductSet:
    #          pos[p] += (10 * (maxl - self.Level[p]), 0)
    #          # node.append(G.add_nodes_from(p))
    #
    #      levels = set(self.Level)
    #      for l in levels:
    #          prodinlevel = [q for q in self.ProductSet if self.Level[q] == l]
    #          prodname = [self.ProductName[q] for q in prodinlevel]
    #          prodname = sorted(prodname)
    #
    #          for k in range(len(prodinlevel)):
    #              ind = self.ProductName.index(prodname[k])
    #              pos[ind] += (0, 100 * k / len(prodinlevel))
    #      edges = G.edges()
    #      colors = [G[u][v]['color'] for u, v in edges]
    #      nx.draw(G, pos, edge_color=colors, node_shape='s',
    #              node_color='b', node_size=500, node_height=300,
    #              width=2,
    #              alpha=0.4)
    #
    #      nx.draw_networkx_labels(G, pos, labels, font_size=12)
    #
    #      plt.axis('off')
    #      return plt

    # This function compute the additional data used in the modes
    # ( indices of the variable,  level in the supply chain, .... )
    def ComputeInstanceData(self):
        self.ComputeIndices()
        self.ComputeLevel()
        self.ComputeMaxLeadTime()
        self.RequieredProduct = [[q for q in self.ProductSet if self.Requirements[q][p] > 0.0] for p in self.ProductSet]
        self.ComputeHasExternalDemand()
        self.ComputeUseForFabrication()
        self.ComputeMaximumArchievableSafetyStock()
        self.MaimumLeadTime = max(self.LeadTimes[p] for p in self.ProductSet)
        self.Delivery = [[0.0 for q in self.ProductSet] for t in self.TimeBucketSet]
        self.ComputePossibleComponent()


    #Compute the possible component for each product (inluding alternative)
    def ComputePossibleComponent(self):

        self.ConsumptionSet = []
        self.PossibleComponents=[[0 for q in self.ProductSet] for p in self.ProductSet]
        for p in self.ProductSet:
            for q in self.ProductSet:
                if self.Requirements[p][q] > 0.0:
                    self.PossibleComponents[p][q] = 1
                    self.ConsumptionSet.append(self.GetConsumptiontuple(q, p))
                elif sum(self.Alternates[c][q] for c in self.ProductSet if self.Requirements[p][c]) >= 1:
                    self.PossibleComponents[p][q] = 1
                    self.ConsumptionSet.append( self.GetConsumptiontuple(q, p))


        self.NrComponent = [sum(1 for q in self.ProductSet
                                if self.PossibleComponents[p][q] > 0.0)
                            for p in self.ProductSet]

        self.NrComponentTotal = sum(self.NrComponent[p] for p in self.ProductSet)

    def GetConsumptiontuple(self, p, q ):
        return (p, q, "%s -> %s" % (p, q))
    # Compute the lead time from a product to its component with the largest sum of lead time
    def ComputeMaxLeadTime(self):
        self.MaxLeadTimeProduct = [0 for p in self.ProductSet]
        levelset = sorted(set(self.Level), reverse=True)
        for l in levelset:
            prodinlevel = [p for p in self.ProductSet if self.Level[p] == l]
            for p in prodinlevel:
                parents = [q for q in self.ProductSet if self.Requirements[p][q] > 0]
                if len(parents) > 0:
                    self.MaxLeadTimeProduct[p] = max([self.MaxLeadTimeProduct[q] for q in parents])
                self.MaxLeadTimeProduct[p] = self.MaxLeadTimeProduct[p] + self.LeadTimes[p]
        self.MaxLeadTime = max(self.MaxLeadTimeProduct[p] for p in self.ProductSet)

    # Return an array giving the maximum amount of time to produce an end-item from each item
    def GetTimeToEnd(self):
        timetoenditem = [0 for p in self.ProductSet]
        levelset = sorted(set(self.Level), reverse=False)
        for l in levelset:
            prodinlevel = [p for p in self.ProductSet if self.Level[p] == l]
            for p in prodinlevel:
                children = [q for q in self.ProductSet if self.Requirements[q][p] > 0]
                if len(children) > 0:
                    timetoenditem[p] = max([timetoenditem[q] for q in children])
                    timetoenditem[p] = max([self.LeadTimes[q] for q in children]) + 1
        return timetoenditem

    # Return the set of descendant of item q
    def GetDescendent(self, q):
        itemtoinvestigate = [q]
        result = [q]
        while len(itemtoinvestigate) > 0:
            nextitemtoinvestigate = []
            for p in itemtoinvestigate:
                children = [q for q in self.ProductSet if self.Requirements[q][p] > 0]
                nextitemtoinvestigate = nextitemtoinvestigate + children
            itemtoinvestigate = nextitemtoinvestigate
            result = result + nextitemtoinvestigate
        return result

    # Fill the array UseForFabrication which is equal to 1 if component p is used to produce r (even not directly)
    def ComputeUseForFabrication(self):
        self.TotalRequirement = [[0 for p in self.ProductSet] for q in self.ProductSet]
        maxlevl = max(self.Level)
        levelset = sorted(set(self.Level), reverse=True)
        for l in levelset:
            prodinlevel = [p for p in self.ProductSet if self.Level[p] == l]
            for p in prodinlevel:
                for q in self.ProductSet:
                    if l == maxlevl:
                        self.TotalRequirement[q][p] = self.Requirements[q][p]
                    else:
                        for c in self.ProductSet:
                            self.TotalRequirement[q][p] = self.TotalRequirement[q][p] \
                                                          + self.Requirements[c][p] * self.TotalRequirement[q][c]

    # This function compute at which level each node is in the supply chain
    def ComputeLevel(self):
        # Maximum lead time and maximum number of level.
        # Get the set of nodes without children
        currentlevelset = [p for p in self.ProductSet
                           if sum(self.Requirements[q][p] for q in self.ProductSet) == 0 \
                           and sum(self.Alternates[q][p] for q in self.ProductSet) == 0
                           ]
        currentlevel = 1
        self.Level = [0 for p in self.ProductSet]
        while len(currentlevelset) > 0:
            nextlevelset = []
            for p in currentlevelset:
                self.Level[p] = currentlevel
                childrenofp = [q for q in self.ProductSet if self.Requirements[p][q] == 1]
                for ch in childrenofp:
                    alternates = [q for q in self.ProductSet if self.Alternates[ch][q] == 1]
                    childrenofp = list(set().union(childrenofp, alternates))
                nextlevelset = nextlevelset + childrenofp
            currentlevelset = set(nextlevelset)
            currentlevel = currentlevel + 1
        self.NrLevel = max(self.Level[p] for p in self.ProductSet)

    # This function set the value of the array HasExternalDemand
    def ComputeHasExternalDemand(self):
        self.HasExternalDemand = [self.YearlyAverageDemand[p] > 0 for p in self.ProductSet]
        self.ProductWithExternalDemand = [p for p in self.ProductSet if self.HasExternalDemand[p]]
        self.ProductWithoutExternalDemand = [p for p in self.ProductSet if not self.HasExternalDemand[p]]

        index = 0
        self.ProductWithExternalDemandIndex = [0 for p in self.ProductSet]
        for p in self.ProductWithExternalDemand:
            self.ProductWithExternalDemandIndex[p] = index
            index = index + 1

        index = 0
        self.ProductWithoutExternalDemandIndex = [0 for p in self.ProductSet]
        for p in self.ProductWithoutExternalDemand:
            self.ProductWithoutExternalDemandIndex[p] = index
            index = index + 1

    # Compute the start of index and the number of variables for the considered instance
    def ComputeIndices(self):
        self.NrTimeBucketWithoutUncertainty = self.NrTimeBucketWithoutUncertaintyAfter + self.NrTimeBucketWithoutUncertaintyBefore
        self.TimeBucketSet = range(self.NrTimeBucket)
        self.ResourceSet = range(self.NrResource)
        self.ProductSet = range(self.NrProduct)

    # This function fills the table MaximumQuanityatT
    def ComputeMaximumArchievableSafetyStock(self):

        self.MaximumQuanityatT = [[0 for p in self.ProductSet] for t in self.TimeBucketSet]

        levelset = sorted(set(self.Level), reverse=True)
        for l in levelset:
            prodinlevel = [p for p in self.ProductSet if self.Level[p] == l]
            for p in prodinlevel:
                for t in self.TimeBucketSet:
                    if t < self.LeadTimes[p]:
                        self.MaximumQuanityatT[t][p] = self.StartingInventories[p]
                    else:
                        RequiredProduct = [q for q in self.ProductSet if self.Requirements[p][q] > 0]
                        if len(RequiredProduct) > 0:
                            minquantity = min(self.MaximumQuanityatT[t - self.LeadTimes[p]][q] for q in RequiredProduct)
                        else:
                            minquantity = Constants.Infinity

                        if self.MaximumQuanityatT[t - 1][p] < Constants.Infinity and minquantity < Constants.Infinity:
                            self.MaximumQuanityatT[t][p] = self.MaximumQuanityatT[t - 1][p] + minquantity
                        else:
                            self.MaximumQuanityatT[t][p] = Constants.Infinity

    # This function compute the dependent demand for components
    def ComputeAverageDemand(self):

        depdemand = [sum(self.ForecastedAverageDemand[t][p] for t in self.TimeBucketSet)
                     / (self.NrTimeBucket - self.NrTimeBucketWithoutUncertaintyBefore) for p in
                     self.ProductSet]

        levelset = sorted(set(self.Level), reverse=False)

        for l in levelset:
            prodinlevel = [p for p in self.ProductSet if self.Level[p] == l]
            for p in prodinlevel:
                depdemand[p] = sum(depdemand[q] * self.Requirements[q][p] for q in self.ProductSet) \
                               + depdemand[p]

        return depdemand

    # This function load the scenario tree from a fil
    # def LoadFromFile(self):
    #     filepath = './Instances/' + self.InstanceName + '_Scenario%s.pkl'%self.ScenarioNr
    #     try:
    #       with open( filepath, 'rb') as input:
    #           result = pickle.load( input )
    #       return result
    #     except:
    #       print "file %r not found" %(filepath)

    # This function read the instance from the file in folder ./Instances/
    def ReadFromFile(self, instancename, distribution,  alternatestructure="Normal", longtimehoizon=False, longtimehorizonperiod = 10, additionaltimehorizon = 0):
        if instancename[0] == "0" or instancename == 'l0"':
            reader = InstanceReaderGrave(self)
        else:
            reader = InstanceReaderTemplemeier(self)
        reader.ReadFromFile(instancename, distribution, longtimehoizon, largetimehorizonperiod=longtimehorizonperiod, alternatetype=alternatestructure, additionaltimehorizon= additionaltimehorizon)


    # Save the scenario tree in a file
    # def SaveCompleteInstanceInFile( self ):
    #    result = None
    #    filepath = '/tmp/thesim/' + self.InstanceName + '_%r.pkl'%self.NrScenarioPerBranch
    #    try:
    #      with open( filepath, 'wb') as output:
    #           pickle.dump(self, output, pickle.HIGHEST_PROTOCOL)
    #    except:
    #      print "file %r not found" %(filepath)

    # Save the Instance in an Excel file
    def SaveCompleteInstanceInExelFile(self):
        # Open the file
        writer = pd.ExcelWriter("./Instances/" + self.InstanceName + ".xlsx", engine='openpyxl')
        # Write generic data about the instance:
        general = [self.InstanceName, self.NrProduct, self.NrTimeBucket, self.NrResource, self.Gamma, self.Distribution,
                   self.NrTimeBucketWithoutUncertaintyBefore, self.NrTimeBucketWithoutUncertaintyAfter]
        columnstab = ["Name", "NrProducts", "NrBuckets", "NrResources", "Gamma", "Distribution",
                      "NrTimeBucketWithoutUncertaintyBefore", "NrTimeBucketWithoutUncertaintyAfter"]
        generaldf = pd.DataFrame(general, index=columnstab)
        generaldf.to_excel(writer, "Generic")
        # Add a tab with an array giving the requirements:
        requirementdf = pd.DataFrame(self.Requirements, index=self.ProductName, columns=self.ProductName)
        requirementdf.to_excel(writer, "Requirement")

        # Add a tab with an array giving the requirements:
        transportcostdf = pd.DataFrame(self.AternateCosts, index=self.ProductName, columns=self.ProductName)
        transportcostdf.to_excel(writer, "AternateCosts")

        # Add a tab with an array giving the alternate sourcing:
        alternaedf = pd.DataFrame(self.Alternates, index=self.ProductName, columns=self.ProductName)
        alternaedf.to_excel(writer, "Alternates")

        # Add a tab with the information of each item:
        productdata = [self.LeadTimes, self.StartingInventories, self.InventoryCosts, self.SetupCosts,
                       self.BackorderCosts, self.YearlyAverageDemand, self.YearlyStandardDevDemands,
                       self.LostSaleCost, self.VariableCost]
        col = ["Leadtimes", "StartingInventories", "InventoryCosts", "SetupCosts", "BackorderCosts", "AverageDemand",
               "StandardDevDemands", "LostSale", "VariableCosts"]
        productdatadf = pd.DataFrame(productdata, columns=self.ProductName, index=col).transpose()
        productdatadf.to_excel(writer, "Productdata")
        # Add some tabs with information on the demand:
        capacitydf = pd.DataFrame(self.ForecastedAverageDemand, index=self.TimeBucketSet, columns=self.ProductName)
        capacitydf.to_excel(writer, "ForecastedAverageDemand")
        capacitydf = pd.DataFrame(self.ForcastedStandardDeviation, index=self.TimeBucketSet, columns=self.ProductName)
        capacitydf.to_excel(writer, "ForcastedStandardDeviation")
        # Add a tab for the capacities and processing times:
        capacitydf = pd.DataFrame(self.Capacity)
        capacitydf.to_excel(writer, "Capacity")
        requirementdf = pd.DataFrame(self.ProcessingTime, index=self.ProductName)
        requirementdf.to_excel(writer, "ProcessingTime")

        writer.save()

        plt = self.DrawSupplyChain()
        #worksheet = writer.sheets['Sheet1']
        plt.savefig("./Temp/testplot.png")

        book = opxl.load_workbook("./Instances/" + self.InstanceName + ".xlsx")
        print ("./Instances/" + self.InstanceName + ".xlsx")
        wb=book.create_sheet("Network")
        img = opxl.drawing.image.Image("./Temp/testplot.png")
        wb.add_image(img)
        book.save("./Instances/" + self.InstanceName + ".xlsx")


    # Read the Instance from an Excel file:
    def ReadInstanceFromExelFile(self, instancename):
        wb2 = opxl.load_workbook("./Instances/" + instancename + ".xlsx")

        # The supplychain is defined in the sheet named "01_LL" and the data are in the sheet "01_SD"
        Genericdf = Tool.ReadDataFrame(wb2, "Generic")
        self.InstanceName = Genericdf.at['Name', 0]
        self.NrProduct = Genericdf.at['NrProducts', 0]
        self.NrTimeBucket = Genericdf.at['NrBuckets', 0]
        self.NrTimeBucketWithoutUncertaintyAfter = Genericdf.at['NrTimeBucketWithoutUncertaintyAfter', 0]
        self.NrTimeBucketWithoutUncertaintyBefore = Genericdf.at['NrTimeBucketWithoutUncertaintyBefore', 0]
        self.NrResource = Genericdf.at['NrResources', 0]
        self.Gamma = Genericdf.at['Gamma', 0]
        self.Distribution = Genericdf.at['Distribution', 0]

        Productdatadf = Tool.ReadDataFrame(wb2, "Productdata")
        self.ProductName = list(Productdatadf.index.values)
        self.LeadTimes = Productdatadf['Leadtimes'].tolist()
        self.InventoryCosts = Productdatadf['InventoryCosts'].tolist()
        self.YearlyAverageDemand = Productdatadf['AverageDemand'].tolist()
        self.YearlyStandardDevDemands = Productdatadf['StandardDevDemands'].tolist()
        self.BackorderCosts = Productdatadf['BackorderCosts'].tolist()
        self.StartingInventories = Productdatadf['StartingInventories'].tolist()
        self.SetupCosts = Productdatadf['SetupCosts'].tolist()
        self.VariableCost = Productdatadf['VariableCosts'].tolist()
        self.LostSaleCost = Productdatadf['LostSale'].tolist()
        self.ComputeIndices()

        Requirementdf = Tool.ReadDataFrame(wb2, "Requirement")
        self.Requirements = [[Requirementdf.at[q, p] for p in self.ProductName] for q in self.ProductName]

        Alternatedf = Tool.ReadDataFrame(wb2, "Alternates")
        self.Alternates = [[Alternatedf.at[q, p] for p in self.ProductName] for q in self.ProductName]
        # self.Alternates = [[0 for p in self.ProductName] for q in self.ProductName]

        # Add a tab with an array giving the requirements:
        # TransportCostdf = Tool.ReadDataFrame(wb2, "TransportCost")
        # self.TransportCost = [[TransportCostdf.at[q, p] for p in self.ProductName] for q in self.ProductName]
        self.AternateCosts = [[0 for p in self.ProductName] for q in self.ProductName]

        Capacitydf = Tool.ReadDataFrame(wb2, "Capacity")
        self.Capacity = [Capacitydf.at[k, 0] for k in self.ResourceSet]

        Processingdf = Tool.ReadDataFrame(wb2, "ProcessingTime")
        self.ProcessingTime = [[Processingdf.at[p, k] for k in self.ResourceSet] for p in self.ProductName]

        forecastedavgdemanddf = Tool.ReadDataFrame(wb2, "ForecastedAverageDemand")
        self.ForecastedAverageDemand = [[forecastedavgdemanddf.at[t, p]
                                         for p in self.ProductName]
                                        for t in self.TimeBucketSet]

        forecastedstddf = Tool.ReadDataFrame(wb2, "ForcastedStandardDeviation")
        self.ForcastedStandardDeviation = [[forecastedstddf.at[t, p]
                                            for p in self.ProductName]
                                           for t in self.TimeBucketSet]

        self.ComputeLevel()
        self.ComputeMaxLeadTime()
        self.ComputeIndices()
        self.ComputeInstanceData()

    def GetLeadTime(self, i, p):
        timetoitem = [0 for q in self.ProductSet]
        child = [i]
        parents = [[] for q in self.ProductSet]
        while True:
            for q in child:
                if len(parents[q]) > 0:
                    timetoitem[q] = min([timetoitem[c] for c in parents[q]])
                timetoitem[q] = timetoitem[q] + self.LeadTimes[q]
                if q == p:
                   return timetoitem[q]

            parents = [[] for q in self.ProductSet]
            nexchild = []
            for current in child:
                for q in self.ProductSet:
                    if self.Requirements[q][current] \
                            or sum(1 for p in self.ProductSet
                                     if self.Alternates[p][current] == 1
                                        and self.Requirements[q][p]) > 0:
                        nexchild.append(q)
                        parents[q].append(current)

            child = nexchild
